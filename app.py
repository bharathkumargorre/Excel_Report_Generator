import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# Read data
df = pd.read_csv("sales_data.csv")

# Calculate total amount
df["Total Amount"] = df["Quantity"] * df["Price"]

# Summaries
category_summary = (
    df.groupby("Category")["Total Amount"]
    .sum()
    .reset_index()
)

monthly_summary = (
    df.groupby("Month")["Total Amount"]
    .sum()
    .reset_index()
)

top_products = (
    df.groupby("Product")["Total Amount"]
    .sum()
    .reset_index()
    .sort_values(by="Total Amount", ascending=False)
)

# Create filename
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"sales_report_{timestamp}.xlsx"

# Write sheets
with pd.ExcelWriter(filename, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Sales Data", index=False)
    category_summary.to_excel(writer, sheet_name="Category Summary", index=False)
    monthly_summary.to_excel(writer, sheet_name="Monthly Summary", index=False)
    top_products.to_excel(writer, sheet_name="Top Products", index=False)

# Load workbook
wb = load_workbook(filename)

# Create Dashboard sheet
dashboard = wb.create_sheet("Dashboard")

dashboard["A1"] = "Sales Dashboard"
dashboard["A3"] = "Total Revenue"
dashboard["B3"] = df["Total Amount"].sum()

dashboard["A5"] = "Total Products"
dashboard["B5"] = len(df["Product"].unique())

dashboard["A7"] = "Total Categories"
dashboard["B7"] = len(df["Category"].unique())

# Styling
header_fill = PatternFill("solid", fgColor="4F81BD")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Format sheets
for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

# Monthly line chart
ws = wb["Monthly Summary"]

data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

line_chart = LineChart()
line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(categories)
line_chart.title = "Monthly Revenue Trend"

ws.add_chart(line_chart, "E2")

# Category pie chart
ws_cat = wb["Category Summary"]

pie_data = Reference(ws_cat, min_col=2, min_row=1, max_row=ws_cat.max_row)
pie_categories = Reference(ws_cat, min_col=1, min_row=2, max_row=ws_cat.max_row)

pie_chart = PieChart()
pie_chart.add_data(pie_data, titles_from_data=True)
pie_chart.set_categories(pie_categories)
pie_chart.title = "Category Distribution"

ws_cat.add_chart(pie_chart, "E2")

wb.save(filename)

print("Dashboard report generated successfully!")